from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import logging


def create_document_xlsx(documents, output_path):
    # print("create_document_xlsx() in toxlsx.py")
    logging.info("create_document_xlsx() in toxlsx.py")

    # print(f"Documents: {documents}")
    # print(f"Documents len: {len(documents)}")

    wb = Workbook()
    ws = wb.active

    # print(f"documents in toxlsx.py: {documents}")
    logging.info(f"documents in toxlsx.py: {documents}")

    # headers = list(documents[0].keys())
    # print("----------------")
    # print(headers)
    # print("----------------")

    # headers = ["_id", "file_path", "transcription", "category", "potential_new_category", "timestamp"]
    headers = ["transcription", "category", "potential_new_category", "file_name", "_id", "timestamp"]

    ws.append(headers)

    column_widths = [75, 25, 25, 20, 20, 20]
    # column_widths = [20, 25, 60, 30, 30, 30]
    for i, column_width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = column_width

    # Define styles
    header_font = Font(bold=True, size=12)  # Font size increased
    standard_font = Font(size=11)  # Standard font for the rest
    header_fill = PatternFill(start_color="20d486", end_color="20d486", fill_type="solid")  

    for cell in ws[1]:  # Apply styles to header row
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    for document in documents:
        # Remove timezone information
        tz_aware_timestamp = document["timestamp"]
        tz_naive_timestamp = tz_aware_timestamp.replace(tzinfo=None)
        
        doc = {
            "_id": document["_id"],
            "file_name": document["file_name"],
            "transcription": document["transcription"],
            "category": document["category"],
            "potential_new_category": document["potential_new_category"],
            "timestamp": tz_naive_timestamp.strftime('%Y-%m-%d %H:%M:%S.%f'),
        }

        ws.append([doc[key] for key in headers])

        for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for c in cell:
                # c.alignment = Alignment(wrap_text=True)
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)


    # Save the workbook
    try:
        wb.save(output_path)
        # print(f"File saved to {output_path}")
        logging.info(f"File saved to {output_path}")
    except Exception as e:
        # print("Error saving the workbook: %s", e)
        logging.error("Error saving the workbook: %s", e)
        raise Exception("Error saving the Excel file")

    return output_path