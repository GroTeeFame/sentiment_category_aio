from typing import List, Dict
from io import BytesIO
import json
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from dotenv import load_dotenv
import os
from azure.ai.textanalytics import TextAnalyticsClient
from azure.core.credentials import AzureKeyCredential
from pydantic import BaseModel
# from contextlib import redirect_stdout
import datetime
import logging

# CONSTANTS ----------------->
BATCH_SIZE = 10
# CONSTANTS ----------------->

###------------------FUNCTIONS------------------------------------>
def add_sender_column_from_excel(file_stream: BytesIO) -> pd.DataFrame:
    """
    Adds a 'Sender' column to a DataFrame extracted from an Excel file, identifying each message's sender
    based on the font color in the conversation column.

    Parameters
    ----------
    file_stream : BytesIO
        A BytesIO stream containing the content of the Excel (.xlsx) file.

    Returns
    -------
    pandas.DataFrame
        A DataFrame with an additional 'Sender' column indicating the sender as 'Customer', 'Bot', or 'Unknown'.

    Raises
    ------
    ValueError
        If the conversation column does not contain color-coded font information.

    Notes
    -----
    The function assumes that the conversation messages are in the fourth column (index 3) of the Excel sheet.
    Adjust the `conversation_column` variable if the messages are in a different column.
    """

    print("add_sender_column_from_excel()")
    logging.info("add_sender_column_from_excel()")
    # Load the workbook from BytesIO object
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
    except Exception as e:
        print(f"ERROR: {e}")
        logging.error(f"ERROR: {e}")
        raise ValueError("Could not read the Excel file stream") from e
    sheet = wb.active
    # Get the headers from the first row
    original_columns = [cell.value for cell in sheet[1]]
    # Convert the worksheet to a pandas DataFrame, dropping the index row
    df = pd.DataFrame(sheet.iter_rows(min_row=2, values_only=True), columns=original_columns)


    if df.iloc[0]['КЛІЄНТ'] == None:
        raise ValueError("Client name have 'None' value. Probably xlsx file is empty...")

    # Index for the conversation messages column
    conversation_column = 3
    # Define the RGB hex colors for customer and bot messages
    customer_color = '135ED6'
    bot_color = '0D8215'

    # Check if conversation_column is within bounds
    if conversation_column >= len(original_columns):
        print("ERROR: The conversation column index is out of bounds.")
        logging.error("ERROR: The conversation column index is out of bounds.")
        raise ValueError("The conversation column index is out of bounds.")

    # List to store sender information
    senders = []

    # Iterate through the rows, getting font color and determining sender
    for row in range(2, len(df) + 2): # Start from the actual data row
        cell = sheet.cell(row=row, column=conversation_column + 1)
        if cell.font and cell.font.color and cell.font.color.type == 'rgb':
            color = cell.font.color.rgb[2:]  # Strip the "00" from the start
        else:
            color = None

        if color == customer_color:
            senders.append('Customer')
        elif color == bot_color:
            senders.append('Bot')
        else:
            senders.append('Unknown')
    # Add the sender information to the DataFrame
    df['Sender'] = senders

    return df


def xlsx_to_json(file_stream: BytesIO) -> List[Dict]:
    """
    Converts an Excel (.xlsx) file into a JSON object.

    Parameters
    ----------
    file_stream : BytesIO
        A BytesIO stream containing the content of the Excel file.

    Returns
    -------
    list
        A list of dictionaries, where each dictionary represents a row from the Excel file.

    Raises
    ------
    ValueError
        If the Excel file cannot be read or is not in a valid format.

    Notes
    -----
    The function reads the entire Excel file and converts it into a JSON object, preserving non-ASCII characters.
    """

    print("xlsx_to_json")
    logging.info("xlsx_to_json")

    # Read the Excel file into a DataFrame
    try:
        excel_data_df = pd.read_excel(file_stream)
    except Exception as e:
        raise ValueError("The Excel file cannot be read or is not in a valid format.") from e

    # Convert the DataFrame to a JSON string with non-ASCII characters preserved
    json_string = excel_data_df.to_json(orient='records', force_ascii=False)

    # Convert the JSON string to a list of dictionaries
    try:
        json_dict = json.loads(json_string)
    except json.JSONDecodeError as e:
        raise ValueError("Error decoding JSON data.") from e

    return json_dict


def convert_json_to_ai_format(file_stream) -> List[List[Dict]]:
    """
    Converts a JSON file into a format suitable for AI processing, structuring the conversation data into documents with metadata.

    Parameters
    ----------
    file_stream : Union[str, file-like object]
        A path to the JSON file or a file-like object containing the JSON data.

    Returns
    -------
    list
        A list of documents, where each document is a list of dictionaries. Each dictionary represents a message with metadata.

    Raises
    ------
    KeyError
        If the expected keys ('КЛІЄНТ', 'Деталі чату', 'Sender') are not present in the JSON data.

    Notes
    -----
    The function processes JSON data extracted from an Excel file. It groups conversation messages into separate documents based on a change in the 'КЛІЄНТ' field.
    """
    
    print("convert_json_to_ai_format")
    logging.info("convert_json_to_ai_format")

    # Load data depending on the input type
    if isinstance(file_stream, list):
        data = file_stream
    else:
        try:
            data = json.load(file_stream)
        except Exception as e:
            raise ValueError("Invalid JSON data or format.") from e

    documents = []
    document = []

    # Iterate over each item in the JSON data
    for i, item in enumerate(data):
        # Verify all required keys exist
        if not all(key in item for key in ('КЛІЄНТ', 'Деталі чату', 'Sender')):
            raise KeyError("Required keys ('КЛІЄНТ', 'Деталі чату', 'Sender') not found in JSON data.")

        # Check for a new conversation based on the 'КЛІЄНТ' key
        if item.get('КЛІЄНТ'):
            if document:
                documents.append(document)
                document = []

        # Handle missing 'Деталі чату' by replacing None or empty with "'"
        chat_details = item.get('Деталі чату') or "'"

        # Create a dictionary for each message with metadata
        temp_dict = {
            'id': i,
            'language': "uk",
            'text': chat_details,
            'speaker': item.get('Sender', 'Unknown')
        }
        document.append(temp_dict)

    # Append the last document if it exists
    if document:
        documents.append(document)

    # print('<-------DOCUMENTS-------')
    # printd(documents)
    # print('-------DOCUMENTS------->')

    return documents


def create_text_analytics_client() -> TextAnalyticsClient:
    """
    Authenticates and creates an instance of the TextAnalyticsClient using environment variables for configuration.

    Returns
    -------
    TextAnalyticsClient
        An authenticated client for interacting with Azure Text Analytics services.

    Raises
    ------
    ValueError
        If `LANGUAGE_ENDPOINT` or `LANGUAGE_KEY` environment variables are not set.

    Notes
    -----
    The function relies on environment variables `LANGUAGE_ENDPOINT` and `LANGUAGE_KEY` for the Azure Text Analytics service configuration.
    Ensure that these variables are correctly set in the environment or a `.env` file.
    """

    print("create_text_analytics_client")

    try:
        # Load environment variables from a .env file
        load_dotenv()

        # Retrieve the endpoint and key for the Text Analytics service
        language_endpoint = os.getenv("LANGUAGE_ENDPOINT")
        language_key = os.getenv("LANGUAGE_KEY")

        # Check if the necessary environment variables are set
        if not language_endpoint or not language_key:
            raise ValueError("Both LANGUAGE_ENDPOINT and LANGUAGE_KEY must be set as environment variables.")

        # Create the Text Analytics client
        ta_credential = AzureKeyCredential(language_key)
        text_analytics_client = TextAnalyticsClient(
            endpoint=language_endpoint, 
            credential=ta_credential
        )
    except Exception as e:
        logging.error(f"ERROR: Can`t create Azure text resource.")
        print(f"ERROR: Can`t create Azure text resource.")

    return text_analytics_client


def analyze_sentiment_with_summary(client: TextAnalyticsClient, documents: List[Dict]) -> List:
    """
    Performs sentiment analysis on a list of documents, calculates average sentiment scores, 
    and determines the overall sentiment for the conversation.

    Parameters
    ----------
    client : TextAnalyticsClient
        An instance of the Azure Text Analytics client used to perform sentiment analysis.
    documents : list
        A list of documents (strings) to analyze for sentiment.

    Returns
    -------
    list
        A list of sentiment analysis results for each document, along with a summary containing
        the overall sentiment and average sentiment scores for the conversation.

    Raises
    ------
    None
    """

    print("analyze_sentiment_with_summary")
    logging.info("analyze_sentiment_with_summary")

    list_of_results = []
    
    # Perform sentiment analysis on the documents
    try:
        print("before p'result = client.analyze_sentiment(documents)'")
        logging.info("before l'result = client.analyze_sentiment(documents)'")
        result = client.analyze_sentiment(documents)
        print("after 'result = client.analyze_sentiment(documents)'")
        logging.info("after 'result = client.analyze_sentiment(documents)'")
    except Exception as e:
        print(f"Sentiment analysis failed: {e}")
        logging.error(f"Sentiment analysis failed: {e}")
        # result = []
        return list_of_results

    # Collect all results in a list
    # list_of_results.append(result)
    list_of_results.extend(result)

    total_positive = 0
    total_neutral = 0
    total_negative = 0
    message_count = 0

    # Aggregate sentiment scores for each document
    for doc in result:
        if not doc.is_error:
            total_positive += doc.confidence_scores.positive
            total_neutral += doc.confidence_scores.neutral
            total_negative += doc.confidence_scores.negative
            message_count += 1

    # Calculate average sentiment scores
    if message_count > 0:
        average_positive = total_positive / message_count
        average_neutral = total_neutral / message_count
        average_negative = total_negative / message_count

        # Determine overall sentiment
        if average_positive > average_negative and average_positive > average_neutral:
            overall_sentiment = "positive"
        elif average_negative > average_positive and average_negative > average_neutral:
            overall_sentiment = "negative"
        else:
            overall_sentiment = "neutral"

        print('------------------------------------------------------------')
        print(f"Av positive : {average_positive:.2f}")
        print(f"Av neutral : {average_neutral:.2f}")
        print(f"Av negative : {average_negative:.2f}")
        print(f"Overall Sentiment for Conversation: {overall_sentiment}")
        print('------------------------------------------------------------')
        logging.info('------------------------------------------------------------')
        logging.info(f"Av positive : {average_positive:.2f}")
        logging.info(f"Av neutral : {average_neutral:.2f}")
        logging.info(f"Av negative : {average_negative:.2f}")
        logging.info(f"Overall Sentiment for Conversation: {overall_sentiment}")
        logging.info('------------------------------------------------------------')

    return list_of_results
    # return result


# Helper function to split documents that longer than 10 into smaller batches:
def split_into_batches(documents: List[Dict], batch_size: int = 10) -> List[List[Dict]]:
    """
    Splits a list of documents into smaller batches of a specified size.

    Parameters
    ----------
    documents : list
        A list of documents to be split into batches.
    batch_size : int, optional
        The maximum number of documents in each batch (default is 10).

    Returns
    -------
    list of lists
        A list containing smaller lists (batches) of documents.

    Raises
    ------
    ValueError
        If `batch_size` is not a positive integer.

    Notes
    -----
    The function divides the input list of documents into sublists (batches) where each sublist has up to `batch_size` documents.
    This is useful for processing large datasets in manageable chunks.
    """

    print("split_into_batches")
    logging.info("split_into_batches")

    # Validate batch size
    if not isinstance(batch_size, int) or batch_size <= 0:
        raise ValueError("batch_size must be a positive integer.")

    batches = []

    # Create batches by slicing the documents list
    for i in range(0, len(documents), batch_size):
        batch = documents[i:i + batch_size]
        batches.append(batch)

    return batches


# function for analysing documents in batches: 
#FIXME: improve with model +
def analyze_sentiment_in_batches(client: TextAnalyticsClient, batches: List[List[Dict]]) -> List:
    """
    Analyzes sentiment for multiple batches of documents using the provided client.

    Parameters
    ----------
    client : TextAnalyticsClient
        An instance of the Azure Text Analytics client used to perform sentiment analysis.
    batches : list of lists
        A list containing batches of documents, where each batch is a list of documents.

    Returns
    -------
    list
        A list of sentiment analysis results for each batch of documents.

    Notes
    -----
    This function processes each batch of documents by calling the `analyze_sentiment_with_summary` function.
    It collects the results from each batch and returns them as a list of results.
    """
    print("-----analyze_sentiment_in_batches-----")
    logging.info("-----analyze_sentiment_in_batches-----")
    batches_result = []

    # Process each batch of documents
    for batch in batches:
        # Analyze sentiment for the current batch and append the result
        result = analyze_sentiment_with_summary(client, batch)
        batches_result.append(result)

    return batches_result

###------------------FUNCTIONS------------------------------------<
class Document(BaseModel): 
    id: int 
    language: str 
    text: str 
    speaker: str

class HTTPException(Exception):
    def __init__(self, status_code: int, detail: str):
        self.status_code = status_code
        self.detail = detail
        super().__init__(f"HTTP Exception {status_code}: {detail}")

# @app.post("/analyze_sentiment")
#FIXME: improve with model +
async def analyze_sentiment(data: List[List[Document]]) -> List:
    """
    Analyzes sentiment of provided documents using Azure Text Analytics service.

    Parameters
    ----------
    data : List[List[Document]]
        A list of lists of Document objects to analyze.

    Returns
    -------
    list
        A list of sentiment analysis results.
    """
    print('/analyze_sentiment')
    logging.info('/analyze_sentiment')
    
    ct = datetime.datetime.now()
    print(f"current time:  {ct}")
    logging.info(f"current time :  {ct}")

    client = create_text_analytics_client()

    result_list = []

    try:
        for document in data:
            if hasattr(document[0], 'dict') and callable(getattr(document[0], 'dict')):
                document_list = [doc.dict() for doc in document]
            else:
                document_list = [doc for doc in document]

            if len(document_list) > 10:
                batches = split_into_batches(document_list, BATCH_SIZE)
                batches_result = analyze_sentiment_in_batches(client, batches)
                result = [item for batch in batches_result for item in batch]
            else:
                result = analyze_sentiment_with_summary(client, document_list)

            result_list.append(result)
    
    except Exception as e:
        logging.error("Error processing document batch: %s", e)
        print("Error processing document batch: %s", e)
        raise HTTPException(status_code=500, detail="Error processing document")

    return result_list


def format_and_save_sentiment_plus(combined_data: dict, output_path: str):
    """
    Processes combined sentiment and original data, appending sentiment scores,
    formatting the data, and saving it to an Excel file.

    Parameters
    ----------
    combined_data : dict
        A dictionary with 'sentiment_data' and 'original_data' lists.
    output_path : str
        The path where the enriched Excel data will be saved.
    """
    logging.info("/format_and_save_sentiment_plus")
    print("/format_and_save_sentiment_plus")

    try:
        sentiment_data = combined_data['sentiment_data']
        original_data = combined_data['original_data']
        
    except KeyError as e:
        logging.error("Combined data missing required key: %s", e)
        raise Exception("Invalid input data structure")

    data_with_sentiment = []
    sentiment_list = [dictionary for sublist in sentiment_data for dictionary in sublist]
    add_scores = lambda scores, sentiment: (
        scores[0] + sentiment['confidence_scores'].positive,
        scores[1] + sentiment['confidence_scores'].neutral,
        scores[2] + sentiment['confidence_scores'].negative
    )
    conversation_scores = [[0, 0, 0, 0] for _ in range(3)]  # conv, cust, asis
    first_message_in_conversation, message_count = 0, 0

    for i, (message, message_sentiment) in enumerate(zip(original_data, sentiment_list)):
        temp = {**message, 'Оцінка повідомлення': message_sentiment['sentiment'],
                'Оцінка клієнта': None, 'Оцінка оператора': None}
        data_with_sentiment.append(temp)

        if message['КЛІЄНТ'] and i != 0:
            for index, scores in enumerate(conversation_scores):
                average_scores = [score / message_count for score in scores[:3]]
                sentiment = max(
                    ('positive', average_scores[0]),
                    ('neutral', average_scores[1]),
                    ('negative', average_scores[2]),
                    key=lambda x: x[1]
                )[0]
                temp_key = ['Оцінка чату', 'Оцінка клієнта', 'Оцінка оператора']
                data_with_sentiment[first_message_in_conversation][temp_key[index]] = sentiment

            conversation_scores = [[0, 0, 0, 0] for _ in range(3)]
            first_message_in_conversation, message_count = i, 0

        conversation_scores[0] = add_scores(conversation_scores[0], message_sentiment)
        if message['Sender'] == 'Customer':
            conversation_scores[1] = add_scores(conversation_scores[1], message_sentiment)
        if message['Sender'] == 'Bot':
            conversation_scores[2] = add_scores(conversation_scores[2], message_sentiment)

        message_count += 1

    # Handle the last conversation
    for index, scores in enumerate(conversation_scores):
        average_scores = [score / message_count for score in scores[:3]]
        sentiment = max(
            ('positive', average_scores[0]),
            ('neutral', average_scores[1]),
            ('negative', average_scores[2]),
            key=lambda x: x[1]
        )[0]
        temp_key = ['Оцінка чату', 'Оцінка клієнта', 'Оцінка оператора']
        data_with_sentiment[first_message_in_conversation][temp_key[index]] = sentiment

    reordered_chat_data = []
    for entry in data_with_sentiment:
        sender_value = entry.pop('Sender', None)
        new_entry = {**entry, 'Повідомлення від': sender_value}
        reordered_chat_data.append(new_entry)

    data_with_sentiment = reordered_chat_data

    # Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    # Set headers
    headers = list(data_with_sentiment[0].keys())
    ws.append(headers)
    column_widths = [15, 7, 20, 60, 20, 10, 20, 20, 10, 18, 14, 15, 16]
    for i, column_width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = column_width

    # Append data rows
    for item in data_with_sentiment:
        ws.append([item[key] for key in headers])
        for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for c in cell:
                c.alignment = Alignment(wrap_text=True)

    fonts = {'Bot': Font(color="FF0D8215"), 'Customer': Font(color="FF135ED6")}
    colors = {
        'positive': PatternFill(start_color='759c7f', fill_type='solid'),
        'neutral': PatternFill(start_color='67b6cf', fill_type='solid'),
        'negative': PatternFill(start_color='c54e50', fill_type='solid'),
        'mixed': PatternFill(start_color='82579c', fill_type='solid')
    }
    top_border = Border(top=Side(style='medium', color='FFD1A93D'))

    ITERATOR_OFFSET = 2

    # Apply formatting
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=ws.max_column, min_row=ITERATOR_OFFSET, max_row=ws.max_row), start=ITERATOR_OFFSET):
        if ws[f"M{i}"].value in fonts:
            ws[f"D{i}"].font = fonts[ws[f"M{i}"].value]

        if ws[f"A{i}"].value:
            for cell in row:
                cell.border = top_border

        for col, key in zip('IJKL', ('I', 'J', 'K', 'L')):
            sentiment = ws[f"{key}{i}"].value
            if sentiment in colors:
                ws[f"{key}{i}"].fill = colors[sentiment]
            else:
                previous_fill = ws[f"{key}{i-1}"].fill
                ws[f"{key}{i}"].fill = PatternFill(start_color=previous_fill.start_color,
                                                   end_color=previous_fill.end_color,
                                                   fill_type=previous_fill.fill_type)

    # Save the workbook
    try:
        wb.save(output_path)
        print(f"File saved to {output_path}")
        logging.info(f"File saved to {output_path}")
    except Exception as e:
        print("Error saving the workbook: %s", e)
        logging.error("Error saving the workbook: %s", e)
        raise Exception("Error saving the Excel file")


async def orchestrate_full_analysis(file_path: str, output_path: str):
    """
    Orchestrates the full analysis process from reading an Excel file to writing an enriched Excel file with sentiment analysis results.

    Parameters
    ----------
    file_path : str
        The path to the uploaded Excel file containing conversation data.
    output_path : str
        The path where the enriched Excel file will be saved.
    """
    print('Starting full analysis process in alp.py')
    logging.info('Starting full analysis process')

    # Read the file
    try:
        with open(file_path, 'rb') as f:
            contents = f.read()
            file_stream = BytesIO(contents)
    except Exception as e:
        print(f"Error reading file: {e}")
        logging.error(f"Error reading file: {e}")
        raise Exception("Could not read file")
    
    # Processing: Adding sender information
    try:
        df = add_sender_column_from_excel(file_stream)
    except Exception as e:
        print(f"ERROR occured (te in orchestrate_full_analysis) : {e}")
        logging.error(f"ERROR occured (te in orchestrate_full_analysis) : {e}")
        raise

    buffer = BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False)
    except Exception as e:
        print(f"!!!ERROR: {e}")
        logging.error(f"!!!ERROR: {e}")
    buffer.seek(0)

    # Convert Excel to JSON
    json_data = xlsx_to_json(buffer)

    # Prepare data for sentiment analysis
    documents = convert_json_to_ai_format(json_data)

    # Analyze sentiment
    analyze_sentiment_result = await analyze_sentiment(documents)

    combined_data = {
        'sentiment_data': analyze_sentiment_result,
        'original_data': json_data
    }

    # Format and save sentiment data
    format_and_save_sentiment_plus(combined_data, output_path)

    # Move or rename final output file
    print(f"Enriched file saved to {output_path}")
    logging.info(f"Enriched file saved to {output_path}")

    return {"status_code": 200}


def printd(text):
    print("-"*50)
    print(text)
    print("-"*50)